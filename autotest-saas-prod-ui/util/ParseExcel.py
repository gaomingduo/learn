# encoding=utf-8
# @time   :2020/5/9 10:23
# @Author :liulaing.song
# @Email  :1056703204@qq.com
# @File   :ParseExcel.py
# @explain:操作excel

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import RED, WHITE
from config.VarConfig import *
import datetime
import time


class ParseExcel(object):
    # 构造函数
    def __init__(self):
        self.workbook = None
        self.excelFile = None

    # 将excel文件加载到内存
    def loadWorkbook(self, excelPathAndName):
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    # 根据sheet名称获取该sheet对象
    def getSheetByName(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    # 获取excel内容
    def getCaseData(self):
        # startTime = datetime.datetime.now()
        startTime = time.strftime('%Y-%m-%d %H:%M:%S')
        # print(startTime)
        paExcel = ParseExcel()
        paExcel.loadWorkbook(casePath)
        sheet = paExcel.getSheetByName('case')
        rows = sheet.max_row

        id = []  # id
        step = []  # 测试步骤
        motion = []  # 动作
        # locationType = []               #操作元素定位方式
        locatorExpression = []  # 操作元素定位表达式
        hoverLocator = []  # 悬浮点击位置定位表达式
        value = []  # 操作值
        getText = []  # 获取文本
        expect = []  # 预期结果
        locatorExpressionResult = []  # 实际结果元素定位表达式
        runResult = []  # 实际结果

        for i in range(2, rows + 1):
            # print(i)
            id.append(sheet.cell(i, 1).value)
            step.append(sheet.cell(i, 2).value)
            motion.append(sheet.cell(i, 3).value)
            # print(motion)
            # locationType.append(sheet.cell(i, 4).value)
            locatorExpression.append(sheet.cell(i, 4).value)
            hoverLocator.append(sheet.cell(i, 5).value)
            value.append(sheet.cell(i, 6).value)
            getText.append(sheet.cell(i, 7).value)
            expect.append(sheet.cell(i, 8).value)
            locatorExpressionResult.append(sheet.cell(i, 9).value)
            runResult.append(sheet.cell(i, 10).value)
        # print(motion)
        return startTime, id, step, motion, locatorExpression, hoverLocator, value, getText, expect, locatorExpressionResult, runResult

    # 填充红颜色
    def fillColorsRed(self, rowNo, colsNo):
        '''

        :param rowNo:行号
        :param colsNo: 列好
        :return: 传入行号和列号 可以填充单元格颜色
        '''
        paExcel = ParseExcel()
        self.workbook = paExcel.loadWorkbook(casePath)
        sheet = paExcel.getSheetByName('case')
        try:
            red = PatternFill(fill_type='solid', fgColor=RED)
            cell = sheet.cell(rowNo, colsNo)
            cell.fill = red
            self.workbook.save(casePath)
        except Exception as e:
            raise e

    # 填充白颜色
    def fillColorsWhite(self, rowNo, colsNo):
        '''

        :param rowNo:行号
        :param colsNo: 列号
        :return: 传入行号和列好 可以填充单元格颜色
        '''
        paExcel = ParseExcel()
        self.workbook = paExcel.loadWorkbook(casePath)
        sheet = paExcel.getSheetByName('case')
        try:
            white = PatternFill(fill_type='solid', fgColor=WHITE)
            cell = sheet.cell(rowNo, colsNo)
            cell.fill = white
            self.workbook.save(casePath)
        except Exception as e:
            raise e

    # 回写内容
    def writeCell(self, rowNo, colsNo, content):
        '''

        :param rowNo: 行号
        :param colsNo: 列好
        :return: 传入行号和列好，对单元格回写返回的内容
        '''
        paExcel = ParseExcel()
        self.workbook = paExcel.loadWorkbook(casePath)
        sheet = paExcel.getSheetByName('case')
        try:
            sheet.cell(row=rowNo, column=colsNo).value = content
            self.workbook.save(casePath)
        except Exception as e:
            raise e


if __name__ == "__main__":
    pe = ParseExcel()
    pe.getCaseData()
